from flask import Flask, request, jsonify
from groq import Groq
from datetime import datetime
import openpyxl
import os
import json
import traceback

app = Flask(__name__)
client = Groq(api_key="PASTE_YOUR_KEY_HERE", timeout=60.0)

EXCEL_FILE = "cimatron_bug_tracker.xlsx"
conversation = []

# ── Setup Excel File ──────────────────────────────────────────────────
def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bug Tracker"
        headers = ["Bug ID","Date","Module","Severity","Summary","Steps to Reproduce","Expected","Actual","Status","Reported By"]
        ws.append(headers)
        from openpyxl.styles import Font, PatternFill, Alignment
        for col, cell in enumerate(ws[1], 1):
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1E3A5F")
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22
        wb.save(EXCEL_FILE)
        print(f"Created {EXCEL_FILE}")

init_excel()

# ── Log Bug to Excel ──────────────────────────────────────────────────
def log_bug(bug):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    bug_id = f"BUG-{str(ws.max_row).zfill(3)}"
    row = [
        bug_id,
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        bug.get("module", "Unknown"),
        bug.get("severity", "Medium"),
        bug.get("summary", ""),
        bug.get("steps", ""),
        bug.get("expected", ""),
        bug.get("actual", ""),
        "Open",
        bug.get("reported_by", "Alex Session")
    ]
    ws.append(row)
    from openpyxl.styles import PatternFill
    severity_colors = {"Critical":"FF444433","High":"FF8C0033","Medium":"FFD70033","Low":"00AA4433"}
    color = severity_colors.get(bug.get("severity","Medium"), "FFD70033")
    for cell in ws[ws.max_row]:
        cell.fill = PatternFill("solid", fgColor=color)
    wb.save(EXCEL_FILE)
    print(f"Bug logged: {bug_id}")
    return bug_id

# ── Routes ────────────────────────────────────────────────────────────
@app.route("/")
def home():
    return open("bug_tracker.html", encoding="utf-8").read()

@app.route("/analyze", methods=["POST"])
def analyze():
    """Analyze bug description and return suggestion WITHOUT logging"""
    global conversation
    try:
        user_msg = request.json.get("message", "")
        conversation.append({"role": "user", "content": user_msg})

        messages = [
            {
                "role": "system",
                "content": """You are Alex, a Senior QA Engineer with 10 years of Cimatron CAD/CAM experience.
Analyze the bug description and respond ONLY with valid JSON:

If it's a bug report:
{
  "type": "bug",
  "summary": "short one-line bug description",
  "module": "one of: NC/Toolpath, Die Design, Mold Design, GPP2, Sheet Metal, General",
  "severity": "your suggestion: Critical/High/Medium/Low",
  "severity_reason": "one sentence explaining why you chose this severity",
  "steps": "numbered steps to reproduce",
  "expected": "what should happen",
  "actual": "what actually happens",
  "message": "brief friendly message to the user"
}

Severity guidelines:
- Critical: crash, data loss, cannot continue work, corrupted file
- High: wrong output, incorrect NC code, major feature broken
- Medium: workaround exists, minor incorrect behavior
- Low: UI glitch, cosmetic issue, minor inconvenience

If it's just a chat/question (not a bug):
{
  "type": "chat",
  "message": "your helpful response"
}

IMPORTANT: Respond with valid JSON only."""
            }
        ] + conversation[-4:]

        res = client.chat.completions.create(
            messages=messages,
            model="llama-3.3-70b-versatile",
            max_tokens=600
        )
        reply = res.choices[0].message.content.strip()
        conversation.append({"role": "assistant", "content": reply})

        try:
            clean = reply.strip()
            if "```" in clean:
                clean = clean.split("```")[1]
                if clean.startswith("json"): clean = clean[4:]
            data = json.loads(clean)
        except:
            return jsonify({"type": "chat", "message": reply})

        return jsonify(data)

    except Exception as e:
        print("ERROR:", traceback.format_exc())
        return jsonify({"type": "chat", "message": f"Error: {str(e)}"})

@app.route("/log", methods=["POST"])
def log():
    """Actually log the bug after user confirms"""
    try:
        bug = request.json
        bug_id = log_bug(bug)
        return jsonify({"success": True, "bug_id": bug_id, "date": datetime.now().strftime("%Y-%m-%d %H:%M")})
    except Exception as e:
        print("LOG ERROR:", traceback.format_exc())
        return jsonify({"success": False, "error": str(e)})

@app.route("/bugs", methods=["GET"])
def get_bugs():
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        bugs = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                bugs.append({"id":row[0],"date":str(row[1]),"module":row[2],"severity":row[3],"summary":row[4],"status":row[8]})
        return jsonify({"bugs": bugs})
    except Exception as e:
        return jsonify({"bugs": [], "error": str(e)})

if __name__ == "__main__":
    print("\n" + "="*45)
    print("  🐛 Alex — Cimatron Bug Tracker")
    print("  🌐 Open: http://127.0.0.1:5002")
    print("  🛑 Stop: Ctrl+C")
    print("="*45 + "\n")
    app.run(port=5002, debug=True)
