from flask import Flask, request, jsonify, send_file
from groq import Groq
from datetime import datetime
import openpyxl
import os
import json
import smtplib
import traceback
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

app = Flask(__name__)
client = Groq(api_key=os.environ.get("GROQ_API_KEY"), timeout=60.0)
print(f"API Key loaded: {bool(os.environ.get('GROQ_API_KEY'))}")

EXCEL_FILE = "chirag_bugs.xlsx"
conversation = []

SYSTEM_PROMPT = """You are Chirag, a Senior QA Engineer with 10 years of experience 
specializing in CAD/CAM software, with deep expertise in Cimatron.

Your background:
- Expert in Cimatron NC, Die Design, Mold Design, GPP2, Sheet Metal modules
- Skilled in testing toolpath generation, fixture design, sheet metal features
- Experience with regression testing, bug reporting, and release validation
- Knowledge of GD&T, CNC G-code verification, and manufacturing workflows

Be direct, professional, friendly, and always use real Cimatron terminology."""

# ── Excel Setup ───────────────────────────────────────────────────────
def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bugs"
        headers = ["Bug ID","Date","Module","Severity","Summary","Steps","Expected","Actual","Status","Reporter"]
        ws.append(headers)
        from openpyxl.styles import Font, PatternFill, Alignment
        for col, cell in enumerate(ws[1], 1):
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0F3460")
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22
        wb.save(EXCEL_FILE)

init_excel()

def log_bug(bug):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    bug_id = f"BUG-{str(ws.max_row).zfill(3)}"
    ws.append([
        bug_id,
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        bug.get("module","Unknown"),
        bug.get("severity","Medium"),
        bug.get("summary",""),
        bug.get("steps",""),
        bug.get("expected",""),
        bug.get("actual",""),
        "Open",
        bug.get("reporter","Chirag Session")
    ])
    from openpyxl.styles import PatternFill
    colors = {"Critical":"FF444433","High":"FF8C0033","Medium":"FFD70033","Low":"00AA4433"}
    c = colors.get(bug.get("severity","Medium"),"FFD70033")
    for cell in ws[ws.max_row]:
        cell.fill = PatternFill("solid", fgColor=c)
    wb.save(EXCEL_FILE)
    return bug_id

def get_bugs():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    bugs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            bugs.append({"id":row[0],"date":str(row[1]),"module":row[2],
                        "severity":row[3],"summary":row[4],"status":row[8]})
    return bugs

# ── Routes ────────────────────────────────────────────────────────────
@app.route("/")
def home():
    return open("command_center.html", encoding="utf-8").read()

@app.route("/chat", methods=["POST"])
def chat():
    global conversation
    try:
        msg = request.json.get("message","")
        conversation.append({"role":"user","content":msg})
        res = client.chat.completions.create(
            messages=[{"role":"system","content":SYSTEM_PROMPT}] + conversation[-8:],
            model="llama-3.3-70b-versatile", max_tokens=800
        )
        reply = res.choices[0].message.content
        conversation.append({"role":"assistant","content":reply})
        return jsonify({"reply": reply})
    except Exception as e:
        return jsonify({"reply": f"Error: {str(e)}"})

@app.route("/analyze", methods=["POST"])
def analyze():
    try:
        msg = request.json.get("message","")
        res = client.chat.completions.create(
            messages=[
                {"role":"system","content":"""You are Chirag, Senior Cimatron QA Engineer.
Analyze the bug and respond ONLY with valid JSON:
{
  "type": "bug",
  "summary": "one-line summary",
  "module": "NC/Toolpath or Die Design or Mold Design or GPP2 or Sheet Metal or General",
  "severity": "Critical or High or Medium or Low",
  "severity_reason": "one sentence why",
  "steps": "numbered steps to reproduce",
  "expected": "expected behavior",
  "actual": "actual behavior",
  "message": "friendly message to user"
}
If it's a question not a bug: {"type":"chat","message":"your answer"}
Respond with valid JSON only."""},
                {"role":"user","content":msg}
            ],
            model="llama-3.3-70b-versatile", max_tokens=600
        )
        reply = res.choices[0].message.content.strip()
        if "```" in reply:
            reply = reply.split("```")[1]
            if reply.startswith("json"): reply = reply[4:]
        return jsonify(json.loads(reply))
    except Exception as e:
        print(traceback.format_exc())
        return jsonify({"type":"chat","message":str(e)})

@app.route("/log", methods=["POST"])
def log():
    try:
        bug = request.json
        bug_id = log_bug(bug)
        return jsonify({"success":True,"bug_id":bug_id,"date":datetime.now().strftime("%Y-%m-%d %H:%M")})
    except Exception as e:
        return jsonify({"success":False,"error":str(e)})

@app.route("/bugs", methods=["GET"])
def bugs():
    try:
        return jsonify({"bugs": get_bugs()})
    except Exception as e:
        return jsonify({"bugs":[],"error":str(e)})

@app.route("/email", methods=["POST"])
def send_email():
    try:
        data = request.json
        sender    = data.get("sender")
        password  = data.get("password")
        recipient = data.get("recipient")
        subject   = data.get("subject","Cimatron Bug Report from Chirag")
        body      = data.get("body","")
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"]    = sender
        msg["To"]      = recipient
        html_body = f"""<html><body style="font-family:sans-serif;background:#0a0e1a;color:#c8d8e8;padding:20px">
        <div style="background:#0f1628;border:1px solid #1e3a5f;border-radius:12px;padding:24px;max-width:600px">
          <h2 style="color:#00d4ff">🔧 Chirag — Cimatron QA Report</h2>
          <p style="color:#4a6080;font-size:12px">{datetime.now().strftime("%Y-%m-%d %H:%M")}</p>
          <hr style="border-color:#1e3a5f"/>
          <pre style="white-space:pre-wrap;color:#c8d8e8;font-size:14px;line-height:1.6">{body}</pre>
          <hr style="border-color:#1e3a5f"/>
          <p style="color:#4a6080;font-size:11px">Sent by Chirag Command Center</p>
        </div></body></html>"""
        msg.attach(MIMEText(body, "plain"))
        msg.attach(MIMEText(html_body, "html"))
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(sender, password)
            server.sendmail(sender, recipient, msg.as_string())
        return jsonify({"success":True,"message":f"Email sent to {recipient}"})
    except Exception as e:
        print(traceback.format_exc())
        return jsonify({"success":False,"error":str(e)})

@app.route("/stats", methods=["GET"])
def stats():
    try:
        bugs = get_bugs()
        modules  = {}
        severity = {"Critical":0,"High":0,"Medium":0,"Low":0}
        for b in bugs:
            m = b.get("module","Unknown")
            modules[m] = modules.get(m,0)+1
            s = b.get("severity","Medium")
            if s in severity: severity[s]+=1
        return jsonify({"total":len(bugs),"open":sum(1 for b in bugs if b.get("status")=="Open"),
                        "critical":severity["Critical"],"modules":modules,"severity":severity,"recent":bugs[-5:][::-1]})
    except Exception as e:
        return jsonify({"error":str(e)})

# ── PWA Routes ────────────────────────────────────────────────────────
@app.route("/manifest.json")
def manifest():
    return send_file("manifest.json", mimetype="application/manifest+json")

@app.route("/service_worker.js")
def sw():
    return send_file("service_worker.js", mimetype="application/javascript")

# ── Run ───────────────────────────────────────────────────────────────
@app.route("/download/bugs")
def download_bugs():
    if os.path.exists(EXCEL_FILE):
        return send_file(EXCEL_FILE, as_attachment=True, download_name="chirag_bugs.xlsx")
    return jsonify({"error": "No bugs file found yet"})


if __name__ == "__main__":
    print("\n" + "="*52)
    print("  🧠 Chirag Command Center")
    print("  🌐 Open: http://127.0.0.1:5000")
    print("  🛑 Stop: Ctrl+C")
    print("="*52 + "\n")
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=False)