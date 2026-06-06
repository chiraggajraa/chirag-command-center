from flask import Flask, request, jsonify
from groq import Groq
from datetime import datetime
from dotenv import load_dotenv
load_dotenv()
import openpyxl
import os
import json
import requests
import base64
import traceback

app = Flask(__name__)
client = Groq(api_key=os.environ.get("GROQ_API_KEY"), timeout=60.0)

# ── ADO Config (replace with your real values) ────────────────────────
ADO_ORG     = "https://dev.azure.com/demo-organization"   # your org URL
ADO_PROJECT = "Cimatron-QA"                               # your project name
ADO_TOKEN   = "PASTE_YOUR_ADO_PAT_HERE"                   # your PAT token
EXCEL_FILE  = "chirag_bugs.xlsx"

# ── ADO Auth Header ───────────────────────────────────────────────────
def get_ado_headers():
    token = base64.b64encode(f":{ADO_TOKEN}".encode()).decode()
    return {
        "Authorization": f"Basic {token}",
        "Content-Type": "application/json-patch+json"
    }

# ── Create ADO Work Item ──────────────────────────────────────────────
def create_ado_work_item(bug):
    url = f"{ADO_ORG}/{ADO_PROJECT}/_apis/wit/workitems/$Bug?api-version=7.0"
    payload = [
        {"op":"add","path":"/fields/System.Title",       "value": f"[{bug.get('id','BUG')}] {bug.get('summary','')}"},
        {"op":"add","path":"/fields/System.Description", "value": f"""
            <b>Module:</b> {bug.get('module','')}<br>
            <b>Severity:</b> {bug.get('severity','')}<br><br>
            <b>Steps to Reproduce:</b><br>{bug.get('steps','')}<br><br>
            <b>Expected:</b> {bug.get('expected','')}<br>
            <b>Actual:</b> {bug.get('actual','')}<br><br>
            <i>Auto-created by Chirag Command Center</i>
        """},
        {"op":"add","path":"/fields/Microsoft.VSTS.Common.Severity", "value": f"2 - {bug.get('severity','Medium')}"},
        {"op":"add","path":"/fields/System.Tags",        "value": f"Cimatron; {bug.get('module','')}; Auto-Logged"},
        {"op":"add","path":"/fields/System.AreaPath",    "value": ADO_PROJECT},
    ]
    res = requests.post(url, headers=get_ado_headers(), json=payload)
    if res.status_code in [200, 201]:
        data = res.json()
        return {"success": True, "id": data["id"], "url": data["_links"]["html"]["href"]}
    else:
        return {"success": False, "error": res.text}

# ── Log bug to Excel ──────────────────────────────────────────────────
def log_bug_excel(bug):
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bugs"
        ws.append(["Bug ID","Date","Module","Severity","Summary","Steps","Expected","Actual","Status","ADO ID","ADO URL"])
        from openpyxl.styles import Font, PatternFill
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0F3460")
        wb.save(EXCEL_FILE)

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    bug_id = f"BUG-{str(ws.max_row).zfill(3)}"
    ws.append([
        bug_id,
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        bug.get("module","Unknown"),
        bug.get("severity","Medium"),
        bug.get("summary",""),
        bug.get("steps",""),
        bug.get("expected",""),
        bug.get("actual",""),
        "Open",
        bug.get("ado_id",""),
        bug.get("ado_url","")
    ])
    from openpyxl.styles import PatternFill
    colors = {"Critical":"FF444433","High":"FF8C0033","Medium":"FFD70033","Low":"00AA4433"}
    c = colors.get(bug.get("severity","Medium"),"FFD70033")
    for cell in ws[ws.max_row]:
        cell.fill = PatternFill("solid", fgColor=c)
    wb.save(EXCEL_FILE)
    return bug_id

# ── Routes ────────────────────────────────────────────────────────────
@app.route("/")
def home():
    return open("ado_integration.html", encoding="utf-8").read()

@app.route("/analyze", methods=["POST"])
def analyze():
    try:
        msg = request.json.get("message","")
        res = client.chat.completions.create(
            messages=[
                {"role":"system","content":"""You are Chirag, Senior Cimatron QA Engineer.
Analyze the bug and respond ONLY with valid JSON:
{
  "type": "bug",
  "summary": "one-line summary",
  "module": "NC/Toolpath or Die Design or Mold Design or GPP2 or Sheet Metal or General",
  "severity": "Critical or High or Medium or Low",
  "severity_reason": "one sentence why",
  "steps": "numbered steps to reproduce",
  "expected": "expected behavior",
  "actual": "actual behavior"
}
If not a bug: {"type":"chat","message":"your answer"}
JSON only."""},
                {"role":"user","content":msg}
            ],
            model="llama-3.3-70b-versatile", max_tokens=500
        )
        reply = res.choices[0].message.content.strip()
        if "```" in reply:
            reply = reply.split("```")[1]
            if reply.startswith("json"): reply = reply[4:]
        return jsonify(json.loads(reply))
    except Exception as e:
        return jsonify({"type":"chat","message":str(e)})

@app.route("/log_and_sync", methods=["POST"])
def log_and_sync():
    try:
        bug = request.json
        result = {"excel": {}, "ado": {}}

        # Log to Excel
        bug_id = log_bug_excel(bug)
        bug["id"] = bug_id
        result["excel"] = {"success": True, "bug_id": bug_id}

        # Create ADO work item
        print(f"Creating ADO work item for {bug_id}...")
        ado_result = create_ado_work_item(bug)
        result["ado"] = ado_result

        if ado_result.get("success"):
            print(f"✅ ADO Work Item #{ado_result['id']} created!")
        else:
            print(f"⚠ ADO failed: {ado_result.get('error','')[:100]}")

        return jsonify(result)
    except Exception as e:
        print(traceback.format_exc())
        return jsonify({"excel":{"success":False},"ado":{"success":False,"error":str(e)}})

@app.route("/bugs", methods=["GET"])
def bugs():
    try:
        if not os.path.exists(EXCEL_FILE):
            return jsonify({"bugs":[]})
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        bugs = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                bugs.append({"id":row[0],"date":str(row[1]),"module":row[2],
                            "severity":row[3],"summary":row[4],"status":row[8],
                            "ado_id":row[9] if len(row)>9 else "","ado_url":row[10] if len(row)>10 else ""})
        return jsonify({"bugs":bugs})
    except Exception as e:
        return jsonify({"bugs":[],"error":str(e)})

if __name__ == "__main__":
    print("\n"+"="*50)
    print("  🔗 Chirag — ADO Integration")
    print(f"  📌 ADO Org: {ADO_ORG}")
    print(f"  📁 Project: {ADO_PROJECT}")
    print("  🌐 Open: http://127.0.0.1:5004")
    print("  🛑 Stop: Ctrl+C")
    print("="*50+"\n")
    app.run(port=5004, debug=True)
