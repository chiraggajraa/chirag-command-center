from flask import Flask, request, jsonify
from groq import Groq
from datetime import datetime
import openpyxl
import os
import json
import traceback

app = Flask(__name__)
client = Groq(api_key="PASTE_YOUR_KEY_HERE", timeout=60.0)

EXCEL_FILE = "multi_agent_test_cases.xlsx"

# ── Setup Excel ───────────────────────────────────────────────────────
def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Cases"
        headers = ["TC ID","Date","Bug Summary","Module","Severity","Test Case Title","Preconditions","Steps","Expected Result","Priority","Status"]
        ws.append(headers)
        from openpyxl.styles import Font, PatternFill, Alignment
        for col, cell in enumerate(ws[1], 1):
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0F3460")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 24
        wb.save(EXCEL_FILE)
        print(f"Created {EXCEL_FILE}")

init_excel()

def save_test_cases(bug_summary, module, severity, test_cases):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    saved_ids = []
    from openpyxl.styles import PatternFill, Alignment
    priority_colors = {"High":"FF8C0033","Medium":"FFD70033","Low":"00AA4433"}

    for i, tc in enumerate(test_cases, 1):
        tc_id = f"TC-{str(ws.max_row).zfill(3)}"
        row = [
            tc_id,
            datetime.now().strftime("%Y-%m-%d %H:%M"),
            bug_summary,
            module,
            severity,
            tc.get("title",""),
            tc.get("preconditions",""),
            tc.get("steps",""),
            tc.get("expected",""),
            tc.get("priority","Medium"),
            "Ready for Testing"
        ]
        ws.append(row)
        color = priority_colors.get(tc.get("priority","Medium"), "FFD70033")
        for cell in ws[ws.max_row]:
            cell.fill = PatternFill("solid", fgColor=color)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        saved_ids.append(tc_id)

    wb.save(EXCEL_FILE)
    return saved_ids

# ── Agent 1: Alex — QA Engineer ───────────────────────────────────────
def alex_analyze(bug_description):
    print("\n🔧 ALEX analyzing bug...")
    res = client.chat.completions.create(
        messages=[
            {
                "role": "system",
                "content": """You are Alex, a Senior QA Engineer with 10 years of Cimatron CAD/CAM experience.
Analyze the bug and respond ONLY with valid JSON:
{
  "summary": "one-line bug summary",
  "module": "one of: NC/Toolpath, Die Design, Mold Design, GPP2, Sheet Metal, General",
  "severity": "Critical/High/Medium/Low",
  "root_cause": "likely root cause in one sentence",
  "test_focus": "what areas Jordan should focus test cases on",
  "instructions": "specific instructions for Jordan on what to test"
}"""
            },
            {"role": "user", "content": f"Analyze this Cimatron bug: {bug_description}"}
        ],
        model="llama-3.3-70b-versatile",
        max_tokens=500
    )
    reply = res.choices[0].message.content.strip()
    try:
        if "```" in reply:
            reply = reply.split("```")[1]
            if reply.startswith("json"): reply = reply[4:]
        return json.loads(reply)
    except:
        return {"summary": bug_description, "module": "General", "severity": "Medium",
                "root_cause": "Unknown", "test_focus": "General testing", "instructions": reply}

# ── Agent 2: Jordan — Test Case Writer ───────────────────────────────
def jordan_write(bug_analysis):
    print("\n🧪 JORDAN writing test cases...")
    res = client.chat.completions.create(
        messages=[
            {
                "role": "system",
                "content": """You are Jordan, a Senior Test Case Writer with 8 years of experience in CAD/CAM QA.
Write detailed test cases based on Alex's bug analysis.
Respond ONLY with valid JSON array of 3 test cases:
[
  {
    "title": "test case title",
    "preconditions": "what must be set up before testing",
    "steps": "numbered steps to execute",
    "expected": "expected result",
    "priority": "High/Medium/Low"
  }
]
Make test cases specific to Cimatron, practical, and thorough."""
            },
            {
                "role": "user",
                "content": f"""Alex analyzed this bug:
Summary: {bug_analysis.get('summary')}
Module: {bug_analysis.get('module')}
Severity: {bug_analysis.get('severity')}
Root cause: {bug_analysis.get('root_cause')}
Test focus: {bug_analysis.get('test_focus')}
Instructions: {bug_analysis.get('instructions')}

Write 3 detailed test cases."""
            }
        ],
        model="llama-3.3-70b-versatile",
        max_tokens=800
    )
    reply = res.choices[0].message.content.strip()
    try:
        if "```" in reply:
            reply = reply.split("```")[1]
            if reply.startswith("json"): reply = reply[4:]
        return json.loads(reply)
    except:
        return [{"title": "Manual Review Required", "preconditions": "See terminal",
                 "steps": reply[:500], "expected": "Bug not reproduced", "priority": "High"}]

# ── Agent 1 Review: Alex reviews Jordan's test cases ─────────────────
def alex_review(bug_analysis, test_cases):
    print("\n🔧 ALEX reviewing Jordan's test cases...")
    res = client.chat.completions.create(
        messages=[
            {
                "role": "system",
                "content": """You are Alex, a Senior QA Engineer reviewing test cases written by Jordan.
Review the test cases and respond ONLY with valid JSON:
{
  "approved": true/false,
  "feedback": "your review feedback in 2-3 sentences",
  "improvements": ["improvement 1", "improvement 2"],
  "overall_score": "score out of 10"
}"""
            },
            {
                "role": "user",
                "content": f"""Review these test cases for bug: {bug_analysis.get('summary')}
Test cases: {json.dumps(test_cases, indent=2)}
Are they thorough enough for a {bug_analysis.get('severity')} severity bug?"""
            }
        ],
        model="llama-3.3-70b-versatile",
        max_tokens=400
    )
    reply = res.choices[0].message.content.strip()
    try:
        if "```" in reply:
            reply = reply.split("```")[1]
            if reply.startswith("json"): reply = reply[4:]
        return json.loads(reply)
    except:
        return {"approved": True, "feedback": reply[:300], "improvements": [], "overall_score": "7/10"}

# ── Agent 2 Refine: Jordan refines based on Alex's review ────────────
def jordan_refine(bug_analysis, test_cases, review):
    print("\n🧪 JORDAN refining test cases based on Alex's review...")
    res = client.chat.completions.create(
        messages=[
            {
                "role": "system",
                "content": """You are Jordan, a Test Case Writer. Refine your test cases based on Alex's review.
Respond ONLY with valid JSON array of 3 improved test cases:
[
  {
    "title": "test case title",
    "preconditions": "setup required",
    "steps": "numbered steps",
    "expected": "expected result",
    "priority": "High/Medium/Low"
  }
]"""
            },
            {
                "role": "user",
                "content": f"""Alex reviewed your test cases for: {bug_analysis.get('summary')}
His feedback: {review.get('feedback')}
Improvements needed: {', '.join(review.get('improvements', []))}
Please refine all 3 test cases incorporating this feedback."""
            }
        ],
        model="llama-3.3-70b-versatile",
        max_tokens=800
    )
    reply = res.choices[0].message.content.strip()
    try:
        if "```" in reply:
            reply = reply.split("```")[1]
            if reply.startswith("json"): reply = reply[4:]
        return json.loads(reply)
    except:
        return test_cases

# ── Routes ────────────────────────────────────────────────────────────
@app.route("/")
def home():
    return open("multi_agent.html", encoding="utf-8").read()

@app.route("/run", methods=["POST"])
def run_pipeline():
    try:
        bug_description = request.json.get("bug", "")
        if not bug_description:
            return jsonify({"error": "No bug description provided"})

        steps = []

        # Step 1: Alex analyzes
        steps.append({"agent": "alex", "action": "analyzing", "message": "Analyzing the bug..."})
        bug_analysis = alex_analyze(bug_description)
        steps.append({
            "agent": "alex",
            "action": "analyzed",
            "message": f"Bug analyzed: {bug_analysis.get('summary')}",
            "data": bug_analysis
        })

        # Step 2: Jordan writes test cases
        steps.append({"agent": "jordan", "action": "writing", "message": "Writing test cases..."})
        test_cases = jordan_write(bug_analysis)
        steps.append({
            "agent": "jordan",
            "action": "written",
            "message": f"Written {len(test_cases)} test cases",
            "data": test_cases
        })

        # Step 3: Alex reviews
        steps.append({"agent": "alex", "action": "reviewing", "message": "Reviewing Jordan's test cases..."})
        review = alex_review(bug_analysis, test_cases)
        steps.append({
            "agent": "alex",
            "action": "reviewed",
            "message": f"Review complete. Score: {review.get('overall_score')}",
            "data": review
        })

        # Step 4: Jordan refines
        steps.append({"agent": "jordan", "action": "refining", "message": "Refining test cases based on feedback..."})
        final_test_cases = jordan_refine(bug_analysis, test_cases, review)
        steps.append({
            "agent": "jordan",
            "action": "refined",
            "message": f"Refined {len(final_test_cases)} test cases ready!",
            "data": final_test_cases
        })

        # Step 5: Save to Excel
        tc_ids = save_test_cases(
            bug_analysis.get("summary"),
            bug_analysis.get("module"),
            bug_analysis.get("severity"),
            final_test_cases
        )
        steps.append({
            "agent": "system",
            "action": "saved",
            "message": f"Saved to Excel: {', '.join(tc_ids)}",
            "data": {"ids": tc_ids, "file": EXCEL_FILE}
        })

        return jsonify({"success": True, "steps": steps, "bug_analysis": bug_analysis, "test_cases": final_test_cases, "review": review})

    except Exception as e:
        print("ERROR:", traceback.format_exc())
        return jsonify({"success": False, "error": str(e)})

if __name__ == "__main__":
    print("\n" + "="*50)
    print("  🤖 Multi-Agent: Alex + Jordan")
    print("  🌐 Open: http://127.0.0.1:5003")
    print("  🛑 Stop: Ctrl+C")
    print("="*50 + "\n")
    app.run(port=5003, debug=True)
