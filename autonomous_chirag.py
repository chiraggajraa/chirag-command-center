from groq import Groq
from apscheduler.schedulers.blocking import BlockingScheduler
from datetime import datetime
import openpyxl
import os
import json
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# ── Config ────────────────────────────────────────────────────────────
GROQ_API_KEY  =  os.environ.get("GROQ_API_KEY")
EMAIL_SENDER  = "chirag.gajraa@gmail.com"           # your gmail
EMAIL_PASS    = "bepb gfkd yalf stqc"    # gmail app password
EMAIL_TO      = "chiraggajra1996@gmail.com"      # who gets the report
EXCEL_FILE    = "chirag_bugs.xlsx"
RUN_EVERY_HOURS = 1                          # how often to run (change to 6, 12, 24 etc)

client = Groq(api_key=GROQ_API_KEY, timeout=60.0)
scheduler = BlockingScheduler()
cycle_count = [0]

# ── Read bugs from Excel ──────────────────────────────────────────────
def get_open_bugs():
    if not os.path.exists(EXCEL_FILE):
        print(f"⚠ No Excel file found: {EXCEL_FILE}")
        return []
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    bugs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[8] == "Open":
            bugs.append({
                "id":       row[0],
                "module":   row[2],
                "severity": row[3],
                "summary":  row[4],
                "steps":    row[5],
                "expected": row[6],
                "actual":   row[7]
            })
    return bugs

# ── Generate test cases for a bug ────────────────────────────────────
def generate_test_cases(bug):
    res = client.chat.completions.create(
        messages=[
            {
                "role": "system",
                "content": """You are Chirag, a Senior QA Engineer with 10 years of Cimatron experience.
Generate 2 test cases for the given bug. Respond ONLY with valid JSON array:
[
  {
    "title": "test case title",
    "steps": "numbered steps",
    "expected": "expected result",
    "priority": "High or Medium or Low"
  }
]"""
            },
            {
                "role": "user",
                "content": f"""Generate test cases for:
Bug ID: {bug['id']}
Summary: {bug['summary']}
Module: {bug['module']}
Severity: {bug['severity']}
Steps: {bug['steps']}
Expected: {bug['expected']}
Actual: {bug['actual']}"""
            }
        ],
        model="llama-3.3-70b-versatile",
        max_tokens=600
    )
    reply = res.choices[0].message.content.strip()
    try:
        if "```" in reply:
            reply = reply.split("```")[1]
            if reply.startswith("json"): reply = reply[4:]
        return json.loads(reply)
    except:
        return [{"title": "Manual review needed", "steps": reply[:200], "expected": "No crash", "priority": "High"}]

# ── Save test cases to Excel ──────────────────────────────────────────
def save_test_cases(bug, test_cases):
    TC_FILE = "chirag_test_cases.xlsx"
    if not os.path.exists(TC_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Cases"
        ws.append(["TC ID","Date","Bug ID","Module","Severity","Title","Steps","Expected","Priority","Status"])
        from openpyxl.styles import Font, PatternFill
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0F3460")
        wb.save(TC_FILE)

    wb = openpyxl.load_workbook(TC_FILE)
    ws = wb.active
    ids = []
    for tc in test_cases:
        tc_id = f"TC-{str(ws.max_row).zfill(3)}"
        ws.append([
            tc_id,
            datetime.now().strftime("%Y-%m-%d %H:%M"),
            bug["id"],
            bug["module"],
            bug["severity"],
            tc.get("title",""),
            tc.get("steps",""),
            tc.get("expected",""),
            tc.get("priority","Medium"),
            "Ready"
        ])
        ids.append(tc_id)
    wb.save(TC_FILE)
    return ids

# ── Send email report ─────────────────────────────────────────────────
def send_report(cycle, bugs_processed, all_test_cases):
    if not EMAIL_SENDER or EMAIL_SENDER == "PASTE_YOUR_GMAIL":
        print("⚠ Email not configured — skipping")
        return

    subject = f"🧠 Chirag Auto QA Report — Cycle #{cycle} — {datetime.now().strftime('%Y-%m-%d %H:%M')}"

    lines = [f"Chirag Autonomous QA Report — Cycle #{cycle}",
             f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
             f"Bugs Processed: {len(bugs_processed)}",
             f"Test Cases Generated: {sum(len(v) for v in all_test_cases.values())}",
             "\n" + "="*50 + "\n"]

    for bug in bugs_processed:
        lines.append(f"[{bug['id']}] {bug['severity']} | {bug['module']}")
        lines.append(f"  {bug['summary']}")
        tcs = all_test_cases.get(bug['id'], [])
        for tc in tcs:
            lines.append(f"  → TC: {tc}")
        lines.append("")

    body = "\n".join(lines)

    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"]    = EMAIL_SENDER
        msg["To"]      = EMAIL_TO
        msg.attach(MIMEText(body, "plain"))
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(EMAIL_SENDER, EMAIL_PASS)
            server.sendmail(EMAIL_SENDER, EMAIL_TO, msg.as_string())
        print(f"✅ Email report sent to {EMAIL_TO}")
    except Exception as e:
        print(f"⚠ Email failed: {e}")

# ── Main QA Cycle ─────────────────────────────────────────────────────
def run_qa_cycle():
    cycle_count[0] += 1
    cycle = cycle_count[0]
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    print(f"\n{'='*55}")
    print(f"  🤖 CHIRAG AUTONOMOUS QA — CYCLE #{cycle}")
    print(f"  🕐 {now}")
    print(f"{'='*55}")

    # Step 1: Read open bugs
    bugs = get_open_bugs()
    print(f"\n📋 Found {len(bugs)} open bugs in Excel")

    if not bugs:
        print("✅ No open bugs — nothing to process this cycle")
        return

    all_test_cases = {}

    # Step 2: Generate test cases for each bug
    for i, bug in enumerate(bugs, 1):
        print(f"\n🔧 [{i}/{len(bugs)}] Processing {bug['id']}: {bug['summary'][:50]}...")
        try:
            test_cases = generate_test_cases(bug)
            tc_ids = save_test_cases(bug, test_cases)
            all_test_cases[bug['id']] = tc_ids
            print(f"  ✅ Generated {len(test_cases)} test cases: {', '.join(tc_ids)}")
        except Exception as e:
            print(f"  ⚠ Error processing {bug['id']}: {e}")

    # Step 3: Send email report
    print(f"\n📧 Sending email report...")
    send_report(cycle, bugs, all_test_cases)

    print(f"\n✅ Cycle #{cycle} complete!")
    print(f"   Bugs processed:      {len(bugs)}")
    print(f"   Test cases created:  {sum(len(v) for v in all_test_cases.values())}")
    print(f"   Next run in:         {RUN_EVERY_HOURS} hour(s)")
    print(f"{'='*55}\n")

# ── Scheduler ─────────────────────────────────────────────────────────
if __name__ == "__main__":
    print(f"\n{'='*55}")
    print(f"  🧠 Chirag Autonomous QA Agent")
    print(f"  📋 Watching: {EXCEL_FILE}")
    print(f"  🕐 Running every {RUN_EVERY_HOURS} hour(s)")
    print(f"  🛑 Stop: Ctrl+C")
    print(f"{'='*55}\n")

    # Run immediately on start
    print("🚀 Running first cycle now...")
    run_qa_cycle()

    # Then schedule
    scheduler.add_job(
        run_qa_cycle,
        'interval',
        hours=RUN_EVERY_HOURS,
        id='qa_cycle',
        next_run_time=None
    )
    scheduler.start()
