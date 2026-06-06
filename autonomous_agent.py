from groq import Groq
from datetime import datetime
from apscheduler.schedulers.blocking import BlockingScheduler
import openpyxl
import os
import json
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

client = Groq(api_key=os.environ.get("GROQ_API_KEY", "PASTE_YOUR_KEY_HERE"), timeout=60.0)

EXCEL_FILE    = "chirag_bugs.xlsx"
REPORT_FILE   = "chirag_autonomous_report.xlsx"
EMAIL_SENDER  = ""   # your gmail
EMAIL_PASS    = ""   # your app password
EMAIL_TO      = ""   # recipient

# ── Read bugs from Excel ──────────────────────────────────────────────
def get_open_bugs():
    if not os.path.exists(EXCEL_FILE):
        print("No bug file found.")
        return []
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    bugs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[8] == "Open":
            bugs.append({"id":row[0],"module":row[2],"severity":row[3],"summary":row[4]})
    return bugs

# ── Generate test cases for a bug ─────────────────────────────────────
def generate_test_cases(bug):
    print(f"  🧠 Generating test cases for {bug['id']}...")
    res = client.chat.completions.create(
        messages=[
            {"role":"system","content":"""You are Chirag, Senior Cimatron QA Engineer.
Generate 2 test cases for the given bug.
Respond ONLY with valid JSON array:
[{"title":"...","steps":"...","expected":"...","priority":"High/Medium/Low"}]"""},
            {"role":"user","content":f"Bug: {bug['summary']}\nModule: {bug['module']}\nSeverity: {bug['severity']}"}
        ],
        model="llama-3.3-70b-versatile",
        max_tokens=500
    )
    reply = res.choices[0].message.content.strip()
    if "```" in reply:
        reply = reply.split("```")[1]
        if reply.startswith("json"): reply = reply[4:]
    return json.loads(reply)

# ── Save test cases to Excel ──────────────────────────────────────────
def save_report(results):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Autonomous QA Report"
    headers = ["Run Date","Bug ID","Bug Summary","Module","Severity","TC Title","Steps","Expected","Priority"]
    ws.append(headers)
    from openpyxl.styles import Font, PatternFill, Alignment
    for col, cell in enumerate(ws[1], 1):
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F3460")
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22

    run_date = datetime.now().strftime("%Y-%m-%d %H:%M")
    for bug, test_cases in results:
        for tc in test_cases:
            ws.append([
                run_date, bug["id"], bug["summary"],
                bug["module"], bug["severity"],
                tc.get("title",""), tc.get("steps",""),
                tc.get("expected",""), tc.get("priority","Medium")
            ])
    wb.save(REPORT_FILE)
    print(f"  💾 Report saved: {REPORT_FILE}")

# ── Send email summary ────────────────────────────────────────────────
def send_email(results, run_time):
    if not EMAIL_SENDER or not EMAIL_TO:
        print("  📧 Email skipped — no credentials set")
        return
    try:
        total_bugs = len(results)
        total_tcs  = sum(len(tcs) for _, tcs in results)

        html = f"""
        <html><body style="font-family:sans-serif;background:#0a0e1a;color:#c8d8e8;padding:20px">
        <div style="background:#0f1628;border:1px solid #1e3a5f;border-radius:12px;padding:24px;max-width:700px">
          <h2 style="color:#00d4ff">🧠 Chirag — Autonomous QA Report</h2>
          <p style="color:#4a6080">Run completed: {run_time}</p>
          <hr style="border-color:#1e3a5f;margin:16px 0"/>
          <div style="display:flex;gap:24px;margin-bottom:16px">
            <div style="text-align:center"><div style="font-size:28px;font-weight:800;color:#00d4ff">{total_bugs}</div><div style="font-size:12px;color:#4a6080">Bugs Processed</div></div>
            <div style="text-align:center"><div style="font-size:28px;font-weight:800;color:#00ff9d">{total_tcs}</div><div style="font-size:12px;color:#4a6080">Test Cases Generated</div></div>
          </div>
          <hr style="border-color:#1e3a5f;margin:16px 0"/>
          {"".join([f'''
          <div style="background:#071420;border:1px solid #1e3a5f;border-radius:8px;padding:12px;margin-bottom:10px">
            <div style="color:#ff6b35;font-weight:700;font-size:13px">{bug["id"]} — {bug["severity"]}</div>
            <div style="color:#c8d8e8;font-size:13px;margin:4px 0">{bug["summary"]}</div>
            <div style="color:#4a6080;font-size:12px">{len(tcs)} test cases generated</div>
          </div>''' for bug, tcs in results])}
          <hr style="border-color:#1e3a5f;margin:16px 0"/>
          <p style="color:#4a6080;font-size:11px">Sent automatically by Chirag Autonomous Agent</p>
        </div></body></html>"""

        msg = MIMEMultipart("alternative")
        msg["Subject"] = f"🧠 Chirag Auto QA Report — {run_time}"
        msg["From"]    = EMAIL_SENDER
        msg["To"]      = EMAIL_TO
        msg.attach(MIMEText(html, "html"))

        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
            s.login(EMAIL_SENDER, EMAIL_PASS)
            s.sendmail(EMAIL_SENDER, EMAIL_TO, msg.as_string())
        print(f"  📧 Email sent to {EMAIL_TO}")
    except Exception as e:
        print(f"  📧 Email failed: {e}")

# ── Main QA Cycle ─────────────────────────────────────────────────────
def run_qa_cycle():
    run_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"\n{'='*52}")
    print(f"  🤖 CHIRAG AUTONOMOUS QA CYCLE")
    print(f"  ⏰ Started: {run_time}")
    print(f"{'='*52}")

    bugs = get_open_bugs()
    if not bugs:
        print("  ✅ No open bugs found — nothing to process!")
        return

    print(f"  🐛 Found {len(bugs)} open bugs to process\n")
    results = []

    for bug in bugs:
        print(f"  Processing {bug['id']}: {bug['summary'][:50]}...")
        try:
            tcs = generate_test_cases(bug)
            results.append((bug, tcs))
            print(f"  ✅ Generated {len(tcs)} test cases")
        except Exception as e:
            print(f"  ❌ Failed for {bug['id']}: {e}")

    if results:
        save_report(results)
        send_email(results, run_time)

    print(f"\n  🏁 Cycle complete — processed {len(results)}/{len(bugs)} bugs")
    print(f"{'='*52}\n")

# ── Scheduler ─────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("""
╔══════════════════════════════════════════════╗
║   🤖 Chirag Autonomous QA Agent              ║
║   Runs automatically every 6 hours           ║
║   Press Ctrl+C to stop                       ║
╚══════════════════════════════════════════════╝
    """)

    # Run immediately on start
    print("▶ Running first cycle now...\n")
    run_qa_cycle()

    # Then schedule every 6 hours
    scheduler = BlockingScheduler()
    scheduler.add_job(
        run_qa_cycle,
        "interval",
        hours=6,
        next_run_time=None,
        id="qa_cycle"
    )

    print("\n⏰ Next run scheduled in 6 hours")
    print("📋 Reports saved to: chirag_autonomous_report.xlsx")
    print("Press Ctrl+C to stop\n")

    try:
        scheduler.start()
    except KeyboardInterrupt:
        print("\n🛑 Autonomous agent stopped.")
